from django.shortcuts import redirect, render
from django.contrib.auth.models import User
from django.contrib import auth
from django.contrib.auth.decorators import login_required
import io
from django.http import FileResponse, JsonResponse
from reportlab.pdfgen import canvas
from .models import *
from PyPDF2 import PdfReader, PdfWriter
import arabic_reshaper
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from bidi.algorithm import get_display
import openpyxl
import datetime
#from datetime import timedelta
from .models import invoice # or Invoice, matching whichever name you pick
from .models import invoice_owner
import logging

logger = logging.getLogger(__name__)  # Add this at the top of your file

"""
objs = building.objects.all()
for i in objs:
    aobjs = apartment.objects.filter(building=i,new_tenant_added=False)
    print(len(aobjs))
    disp = 0
    for j in aobjs:
        j.display_order = disp
        j.save()
        disp += 1


objs = apartment.objects.all()
for i in objs:
    i.payment_method = "1"
    i.annual_rent = "1"
    i.save()
"""


def get_user_profile(userobj):
    try:
        obj = user_profile.objects.get(user=userobj)
    except:
        obj = user_profile()
        obj.user = userobj
        obj.save()
    return obj

def login(request):
    if request.user.is_authenticated:
        return redirect("/home")
    template = "login.html"
    context = {}
    if request.method == "POST":
        user = User.objects.filter(username=request.POST['user'])
        if (len(user) > 0):
            user = user[0]
            if (user.check_password(request.POST['pass'])):
                auth.login(request,user)
                return redirect("/home")
            else:
                context['message'] = "The entered password is not correct!"
        else:
            context['message'] = "The entered user was not found!"
    return render(request,template,context)

def logout(request):
    auth.logout(request)
    return redirect("/")

@login_required(login_url='/')
def home(request):
    template = "home.html"
    context = {}
    obj = get_user_profile(request.user)

    context['type_of_user'] = obj.type_of_user == 'd'
    if obj.type_of_user in ('v','w'):
        temp = building.objects.all()
        temp2 = invoice_owner.objects.all()
        temp_list = []
        temp2_list = []
        for i in temp:
            if i.owner in obj.invoice_owner_allowed.all():
                temp_list.append(i)
        for i in temp2:
            if i in obj.invoice_owner_allowed.all():
                temp2_list.append(i)
        context['objs'] = temp_list
        context['owners'] = temp2_list
    else:
        context['objs'] = building.objects.all()
        context['owners'] = invoice_owner.objects.all()
    return render(request,template,context)

@login_required(login_url='/')
def delete_building(request,id):
    obj = building.objects.get(pk=id)
    obj.delete()
    return redirect("/home")

@login_required(login_url='/')
def building_form(request):
    template = "new_building_form.html"
    context = {}
    context['objs'] = invoice_owner.objects.all()
    if request.method == "POST":
        if request.POST['name'] and request.POST['invoice-owner']:
            objs = building.objects.filter(name=request.POST['name'])
            if len(objs) == 0:
                obj = building()
                obj.name = request.POST['name']
                obj.owner = invoice_owner.objects.get(pk=request.POST['invoice-owner'])
                obj.save()
                return redirect('/home')
            else:
                context['message'] = "Building with that name already exists."
        else:
            context['message'] = "Entered data is not valid."
    return render(request,template,context)

@login_required(login_url='/')
def apartments(request,id):
    if id:
        template = "apartments.html"
        context = {}
        obj = get_user_profile(request.user)
        context['type_of_user'] = obj.type_of_user == 'd'
        building_obj = building.objects.get(pk=id)
        objs = apartment.objects.filter(building=building_obj,temp_del=False,new_tenant_added=False).order_by("display_order")
        del_objs = apartment.objects.filter(building=building_obj,temp_del=True,new_tenant_added=False)
        temp_del_objs = []
        for i in del_objs:
            day_diff = datetime.date.today() - i.temp_del_date
            if day_diff.days >= 7:
                i.delete()
            else:
                temp_del_objs.append(i)
        context['bobj'] = building_obj
        context['objs'] = objs
        context['del_objs'] = temp_del_objs
        if context['type_of_user']:
            context['owners'] = invoice_owner.objects.all()
        else:
            temp2 = invoice_owner.objects.all()
            temp2_list = []
            for i in temp2:
                if i in obj.invoice_owner_allowed.all():
                    temp2_list.append(i)
            context['owners'] = temp2_list
        if request.method == "POST":
            if request.POST['name'] and request.POST['invoice-owner']:
                objs = building.objects.filter(name=request.POST['name'])
                if (len(objs) == 0 and building_obj.name != request.POST['name']) or (len(objs) == 1 and building_obj.name == request.POST['name']):
                    building_obj.name = request.POST['name']
                    building_obj.owner = invoice_owner.objects.get(pk=request.POST['invoice-owner'])
                    building_obj.save()
                    return redirect("/apartments/{}".format(id))
                else:
                    context['message'] = "Building with that name already exists."
            else:
                context['message'] = "Entered data is not valid."
        return render(request,template,context)
    else:
        return redirect('/home')

@login_required(login_url='/')
def apartment_form(request,id,prev_id):
    if id:
        template = "new_apartment_form.html"
        context = {'id':id}
        if prev_id != "x":
            aobj = apartment.objects.get(pk=int(prev_id))
            context['aobj'] = aobj
        if request.method == "POST":
            if request.POST['name'] and request.POST['num'] and request.POST['phone'] and request.POST['type_of'] and request.POST['dob'] and request.POST['cnum'] and request.POST['enum'] and request.POST['rent'] and request.POST['payment_method']:
                bobj = building.objects.get(pk=id)
                obj = apartment()
                obj.aprt_number = request.POST['num']
                obj.name = request.POST['name']
                obj.type_of = request.POST['type_of']
                obj.dob = request.POST['dob']
                obj.phone_nmber = request.POST['phone']
                obj.elect_number = request.POST['enum']
                obj.contract_number = request.POST['cnum']
                obj.annual_rent = request.POST['rent']
                obj.payment_method = request.POST['payment_method']
                obj.note = request.POST['note']
                obj.building = bobj
                obj.save()
                if prev_id == "x":
                    linkobj = tenant_link()
                    linkobj.save()
                    temp_objs = apartment.objects.filter(building=bobj).order_by("-display_order")
                    if (len(temp_objs) > 0):
                        obj.display_order = temp_objs[0].display_order
                else:
                    obj.display_order = aobj.display_order
                    aobj.new_tenant_added = True
                    aobj.save()
                    linkobj = aobj.aprt_link
                    if (None == linkobj):
                        tempobj = tenant_link()
                        tempobj.save()
                        tempobj.apartments.add(aobj)
                        tempobj.save()
                        aobj.aprt_link = tempobj
                        aobj.save()
                        linkobj = aobj.aprt_link

                linkobj.apartments.add(obj)
                linkobj.save()
                obj.aprt_link = linkobj
                obj.save()
                return redirect("/apartments/{}".format(id))
            else:
                context['message'] = "Entered data is not valid."
        return render(request,template,context)
    else:
        return redirect("/home")

@login_required(login_url='/')
def edit_apartment_form(request,id):
    if id:
        template = "apartments_invoice.html"
        context = {}
        if request.method == "POST":
            if request.POST['name'] and request.POST['num'] and request.POST['phone'] and request.POST['type_of'] and request.POST['dob'] and request.POST['cnum'] and request.POST['enum'] and request.POST['rent'] and request.POST['payment_method']:
                objs = apartment.objects.filter(pk=id)
                obj = objs[0]
                obj.aprt_number = request.POST['num']
                obj.name = request.POST['name']
                obj.type_of = request.POST['type_of']
                obj.dob = request.POST['dob']
                obj.phone_nmber = request.POST['phone']
                obj.elect_number = request.POST['enum']
                obj.contract_number = request.POST['cnum']
                obj.annual_rent = request.POST['rent']
                obj.payment_method = request.POST['payment_method']
                obj.note = request.POST['note']
                obj.save()
                return redirect("/invoices/{}".format(id))
            else:
                context['message'] = "Entered data is not valid."
        obj = apartment.objects.filter(pk=id)[0]
        objs = invoice.objects.filter(is_deleted=False,apartment=obj)
        context['aobj'] = obj
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect("/home")

@login_required(login_url='/')
def delete_apartment(request,id):
    if id:
        obj = apartment.objects.filter(pk=id)[0]
        if obj.temp_del:
            obj.temp_del = False
        else:
            obj.temp_del = True
            obj.temp_del_date = datetime.date.today()
        obj.save()
        bid = obj.building.id
        return redirect("/apartments/{}".format(bid))
    else:
        return redirect("/home")

@login_required(login_url='/')
def invoices(request,id):
    if id:
        template = "apartments_invoice.html"
        context = {}
        obj = get_user_profile(request.user)

        context['type_of_user'] = obj.type_of_user == 'd'
        context['write_priv'] = obj.type_of_user == 'w'
        aobj = apartment.objects.get(pk=id)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=False).order_by("today_date")
            else:
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=False).order_by("-today_date")
        else:
            objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=False)
        context['aobj'] = aobj
        context['date_dis'] = aobj.dob.strftime("%Y-%m-%d")
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect('/home')

@login_required(login_url='/')
def other_invoices(request,id):
    if id:
        template = "apartments_other_invoice.html"
        context = {}
        obj = get_user_profile(request.user)

        context['type_of_user'] = obj.type_of_user == 'd'
        context['write_priv'] = obj.type_of_user == 'w'
        aobj = apartment.objects.get(pk=id)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=True).order_by("today_date")
            else:
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=True).order_by("-today_date")
        else:
            objs = invoice.objects.filter(is_deleted=False,apartment=aobj,other_payment=True)
        context['aobj'] = aobj
        context['date_dis'] = aobj.dob.strftime("%Y-%m-%d")
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect('/home')

@login_required(login_url='/')
def invoice_form(request,id):
    if id:
        template = "new_invoice_form.html"
        context = {'id':id}
        objs = reversed(list(invoice.objects.filter(is_deleted=False,apartment=apartment.objects.get(pk=id),other_payment=False)))
        temp = []
        temp1 = []
        tempcount = 0
        for i in objs:
            temp.append(i.remaining_amount)
            temp1.append("From : {} To : {}".format(i.from_date,i.to_date))
            tempcount += 1
            if tempcount == 3:
                break
        context['prev_trans'] = zip(temp1,temp)
        if request.method == "POST":
            if request.POST['amount'] and (request.POST['payment'] == "Cash" or (request.POST['payment'] == "Transfer" and request.POST['bank'] and request.POST['trans_date'])) and request.POST['fdate'] and request.POST['tdate']:
                obj = invoice()
                obj.user = request.user
                obj.apartment = apartment.objects.get(pk=id)
                obj.owner = apartment.objects.get(pk=id).building.owner
                obj.amount = request.POST['amount']
                if request.POST['ramount']:
                    obj.remaining_amount = request.POST['ramount']
                else:
                    obj.remaining_amount = 0
                obj.payment_method = request.POST['payment']
                if (request.POST['payment'] == "Transfer"):
                    obj.bank_of_transfer = request.POST['bank']
                    obj.transfer_date = request.POST['trans_date']
                obj.from_date = request.POST['fdate']
                obj.to_date = request.POST['tdate']
                if request.POST['note']:
                    obj.note = request.POST['note']
                temp_len_inv = len(invoice.objects.filter(is_deleted=False,owner=obj.owner))
                if temp_len_inv > 0:
                    obj.invoice_number = invoice.objects.filter(is_deleted=False,owner=obj.owner).order_by("-invoice_number")[0].invoice_number + 1
                obj.save()
                return redirect("/invoices/{}".format(id))
            else:
                context['message'] = "Entered data is not valid."
        return render(request,template,context)
    else:
        return redirect("/home")

@login_required(login_url='/')
def other_invoice_form(request,id):
    if id:
        template = "new_other_invoice_form.html"
        context = {'id':id}
        objs = reversed(list(invoice.objects.filter(is_deleted=False,apartment=apartment.objects.get(pk=id),other_payment=True)))
        temp = []
        temp1 = []
        tempcount = 0
        for i in objs:
            temp.append(i.remaining_amount)
            temp1.append("From : {} To : {}".format(i.from_date,i.to_date))
            tempcount += 1
            if tempcount == 3:
                break
        context['prev_trans'] = zip(temp1,temp)
        if request.method == "POST":
            if request.POST['amount'] and (request.POST['payment'] == "Cash" or (request.POST['payment'] == "Transfer" and request.POST['bank'] and request.POST['trans_date'])) and request.POST['fdate'] and request.POST['tdate'] and request.POST['type']:
                obj = invoice()
                obj.user = request.user
                obj.apartment = apartment.objects.get(pk=id)
                obj.owner = apartment.objects.get(pk=id).building.owner
                obj.amount = request.POST['amount']
                if request.POST['ramount']:
                    obj.remaining_amount = request.POST['ramount']
                else:
                    obj.remaining_amount = 0
                obj.payment_method = request.POST['payment']
                if (request.POST['payment'] == "Transfer"):
                    obj.bank_of_transfer = request.POST['bank']
                    obj.transfer_date = request.POST['trans_date']
                obj.from_date = request.POST['fdate']
                obj.to_date = request.POST['tdate']
                if request.POST['note']:
                    obj.note = request.POST['note']
                temp_len_inv = len(invoice.objects.filter(is_deleted=False,owner=obj.owner))
                if temp_len_inv > 0:
                    obj.invoice_number = invoice.objects.filter(is_deleted=False,owner=obj.owner).order_by("-invoice_number")[0].invoice_number + 1
                obj.other_payment = True
                if request.POST['type'] == "Other":
                    obj.payment_type = request.POST['type-other']
                else:
                    obj.payment_type = request.POST['type']
                obj.save()
                return redirect("/other-invoices/{}".format(id))
            else:
                context['message'] = "Entered data is not valid."
        return render(request,template,context)
    else:
        return redirect("/home")

def get_reversed(a):
    retval = ""
    for i in range(1,len(a)):
        retval += a[len(a)-i]
    retval += a[0]
    return retval

@login_required(login_url='/')
def print_invoice(request,id):
    aprt_types = {"Apartment":"شقة","Floor":"دور","Home":"أرض","Store":"محل ","Studio":"ملحق"}
    invoice_name = "Invoice.pdf"

    pdfmetrics.registerFont(TTFont("Arabic","font.ttf"))

    obj = invoice.objects.filter(pk=id)

    buffer = io.BytesIO()

    p = canvas.Canvas(buffer)
    p.setFont('Arabic', 13)

    offset = 402
    remain_offset = -17
    x_remain_offset = 0

    if (len(obj) == 1):

        if (obj[0].payment_method == "Cash"):

            if (obj[0].remaining_amount > 0):
                invoice_name = "Cash_Remain.pdf"
                remain_offset = 35
                offset = offset + 25
                p.drawString(270, 410+remain_offset, "{}".format(obj[0].remaining_amount))
                p.drawString(270, 410-offset+remain_offset, "{}".format(obj[0].remaining_amount))

            p.drawString(340, 523+remain_offset, "{}".format(obj[0].from_date.year))
            p.drawString(374, 523+remain_offset, "{}".format(obj[0].from_date.month))
            p.drawString(392, 523+remain_offset, "{}".format(obj[0].from_date.day))
            p.drawString(338, 523-offset+remain_offset, "{}".format(obj[0].from_date.year))
            p.drawString(372, 523-offset+remain_offset, "{}".format(obj[0].from_date.month))
            p.drawString(390, 523-offset+remain_offset, "{}".format(obj[0].from_date.day))

            p.drawString(145, 523+remain_offset, "{}".format(obj[0].to_date.year))
            p.drawString(180, 523+remain_offset, "{}".format(obj[0].to_date.month))
            p.drawString(200, 523+remain_offset, "{}".format(obj[0].to_date.day))
            p.drawString(143, 523-offset+remain_offset, "{}".format(obj[0].to_date.year))
            p.drawString(178, 523-offset+remain_offset, "{}".format(obj[0].to_date.month))
            p.drawString(198, 523-offset+remain_offset, "{}".format(obj[0].to_date.day))

            p.drawString(270, 485+remain_offset, get_display(arabic_reshaper.reshape("نقدا")))
            p.drawString(270, 485-offset+remain_offset, get_display(arabic_reshaper.reshape("نقدا")))

            p.drawString(45, 760-offset+remain_offset, "{}".format(obj[0].today_date.strftime("%Y/%m/%d")))
            p.drawString(15, 740-offset+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].user.username))))
            p.drawString(380, 760-offset+remain_offset, "{:05d}".format(obj[0].invoice_number))
            p.drawString(360, 740-offset+remain_offset, obj[0].apartment.contract_number)
            p.drawString(270, 720-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].apartment.name)) )
            p.drawString(270, 680-offset+remain_offset, "{}".format(obj[0].amount))
            if obj[0].other_payment:
                p.drawString(260, 640-offset+remain_offset, obj[0].payment_type)
            else:
                p.drawString(260, 640-offset+remain_offset, get_display(arabic_reshaper.reshape(aprt_types[obj[0].apartment.type_of])))
            p.drawString(280, 600-offset+remain_offset, "{}".format(obj[0].apartment.aprt_number))
            p.drawString(270, 560-offset+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].apartment.building.name))))
            p.setFont('Arabic', 10)
            p.drawString(110, 455-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].note)))
            p.setFont('Arabic', 13)

        else:
            remain_offset = -13.5
            x_remain_offset = 0
            if (obj[0].remaining_amount > 0):
                invoice_name = "Transfer_Remain2.pdf"

                p.drawString(270, 410+remain_offset, "{}".format(obj[0].remaining_amount))
                #p.drawString(270, 410-offset+remain_offset, "{}".format(obj[0].remaining_amount))

                p.drawString(135+x_remain_offset, 472+remain_offset, "{}".format(obj[0].transfer_date.year))
                p.drawString(175+x_remain_offset, 472+remain_offset, "{}".format(obj[0].transfer_date.month))
                p.drawString(202+x_remain_offset, 472+remain_offset, "{}".format(obj[0].transfer_date.day))
                #p.drawString(135-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.year))
                #p.drawString(175-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.month))
                #p.drawString(202-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.day))

                x_remain_offset = 0

                p.setFont('Arabic', 10)

                p.drawString(255, 479+remain_offset, get_display(arabic_reshaper.reshape(obj[0].bank_of_transfer)))
                #p.drawString(255, 480-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].bank_of_transfer)))

                p.setFont('Arabic', 13)
            else:
                invoice_name = "Transfer_Invoice2.pdf"

                p.drawString(135-x_remain_offset, 475+remain_offset, "{}".format(obj[0].transfer_date.year))
                p.drawString(175-x_remain_offset, 475+remain_offset, "{}".format(obj[0].transfer_date.month))
                p.drawString(202-x_remain_offset, 475+remain_offset, "{}".format(obj[0].transfer_date.day))
                #p.drawString(135-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.year))
                #p.drawString(175-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.month))
                #p.drawString(202-x_remain_offset, 482-offset+remain_offset, "{}".format(obj[0].transfer_date.day))

                p.setFont('Arabic', 10)

                p.drawString(255, 475+remain_offset, get_display(arabic_reshaper.reshape(obj[0].bank_of_transfer)))
                #p.drawString(255, 480-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].bank_of_transfer)))

                p.setFont('Arabic', 13)

            p.drawString(340-x_remain_offset, 520+remain_offset, "{}".format(obj[0].from_date.year))
            p.drawString(374-x_remain_offset, 520+remain_offset, "{}".format(obj[0].from_date.month))
            p.drawString(392-x_remain_offset, 520+remain_offset, "{}".format(obj[0].from_date.day))
            #p.drawString(338-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].from_date.year))
            #p.drawString(372-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].from_date.month))
            #p.drawString(390-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].from_date.day))

            p.drawString(145-x_remain_offset, 520+remain_offset, "{}".format(obj[0].to_date.year))
            p.drawString(180-x_remain_offset, 520+remain_offset, "{}".format(obj[0].to_date.month))
            p.drawString(200-x_remain_offset, 520+remain_offset, "{}".format(obj[0].to_date.day))
            #p.drawString(143-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].to_date.year))
            #p.drawString(178-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].to_date.month))
            #p.drawString(198-x_remain_offset, 525-offset+remain_offset, "{}".format(obj[0].to_date.day))

        p.drawString(45, 760+remain_offset, "{}".format(obj[0].today_date.strftime("%Y/%m/%d")))
        #p.drawString(45, 760-offset+remain_offset, "{}".format(obj[0].today_date.strftime("%Y/%m/%d")))

        p.drawString(15, 740+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].user.username))))
        #p.drawString(15, 740-offset+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].user.username))))

        p.drawString(380, 760+remain_offset, "{:05d}".format(obj[0].invoice_number))
        #p.drawString(380, 760-offset+remain_offset, "{:05d}".format(obj[0].invoice_number))

        p.drawString(380, 740+remain_offset, obj[0].apartment.contract_number)
        #p.drawString(380, 740-offset+remain_offset, obj[0].apartment.contract_number)

        p.drawString(270, 720+remain_offset, get_display(arabic_reshaper.reshape(obj[0].apartment.name)) )
        #p.drawString(270, 720-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].apartment.name)) )

        p.drawString(270, 680+remain_offset, "{}".format(obj[0].amount))
        #p.drawString(270, 680-offset+remain_offset, "{}".format(obj[0].amount))

        if obj[0].other_payment:
            p.drawString(260, 640+remain_offset, obj[0].payment_type)
        else:
            p.drawString(260, 640+remain_offset, get_display(arabic_reshaper.reshape(aprt_types[obj[0].apartment.type_of])))
        #p.drawString(260, 640-offset+remain_offset, get_display(arabic_reshaper.reshape(aprt_types[obj[0].apartment.type_of])))

        p.drawString(280, 600+remain_offset, "{}".format(obj[0].apartment.aprt_number))
        #p.drawString(280, 600-offset+remain_offset, "{}".format(obj[0].apartment.aprt_number))

        p.drawString(270, 560+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].apartment.building.name))))
        #p.drawString(270, 560-offset+remain_offset, "{}".format(get_display(arabic_reshaper.reshape(obj[0].apartment.building.name))))

        p.setFont('Arabic', 10)

        p.drawString(110, 455+remain_offset, get_display(arabic_reshaper.reshape(obj[0].note)))
        #p.drawString(110, 455-offset+remain_offset, get_display(arabic_reshaper.reshape(obj[0].note)))


    p.showPage()
    p.save()

    buffer.seek(0)

    new_pdf = PdfReader(buffer)

    existing_pdf = PdfReader(invoice_name)
    output = PdfWriter()

    page = existing_pdf.pages[0]
    page.merge_page(new_pdf.pages[0])
    output.add_page(page)

    outputStream = io.BytesIO()
    output.write(outputStream)
    outputStream.seek(0)

    return FileResponse(outputStream , as_attachment=True, filename='invoice.pdf')

@login_required(login_url='/')
def delete_invoice(request,id):
    obj = invoice.objects.get(pk=id)
    try:
        tempId = obj.apartment.id
        obj.is_deleted = True
        obj.save()
        return redirect("/invoices/{}".format(tempId))
    except:
        obj.is_deleted = True
        obj.save()
        return redirect('/home')

@login_required(login_url='/')
def actual_delete_invoice(request,id):
    obj = invoice.objects.get(pk=id)
    try:
        tempId = obj.apartment.id
        obj.delete()
        return redirect("/deleted-invoices")
    except:
        obj.delete()
        return redirect('/home')


from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib.auth.models import User
#from datetime import datetime
from .models import invoice, invoice_owner



@login_required(login_url='/')
def owner_invoices(request, id):
    """
    A single view that:
      - Takes 'id' from the URL ("x" or an integer).
      - Builds a base queryset of invoices, respecting user ownership restrictions.
      - Uses advanced GET-based filters if present. Otherwise, uses your old POST-based sort.
      - Paginates the results.
      - Preserves filters in pagination links.
    """
    # 1) If no ID is provided, redirect to home
    if not id:
        return redirect('/home')

    template = "owners_invoice.html"
    context = {}
    obj = get_user_profile(request.user)
    context['type_of_user'] = (obj.type_of_user == 'd')

    # 2) Keep track of which owner is selected in the URL
    if id != "x":
        context['sel_owner'] = int(id)

    # ---------------------------------------------
    # 3) Build the BASE invoice queryset
    # ---------------------------------------------
    if id != "x":
        base_qs = invoice.objects.filter(is_deleted=False, owner__id=int(id))
    else:
        # If "x" => All owners, but if user is 'v' or 'w', only show allowed owners
        base_qs = invoice.objects.filter(is_deleted=False)
        if obj.type_of_user in ('v', 'w'):
            allowed_ids = obj.invoice_owner_allowed.all().values_list('id', flat=True)
            base_qs = base_qs.filter(owner__id__in=allowed_ids)

    # ---------------------------------------------
    # 4) Read GET filter parameters
    # ---------------------------------------------
    invoice_number  = request.GET.get('invoice_number', '').strip()
    owner_id_filter = request.GET.get('owner_id', '').strip()  # from advanced filter form
    user_id_filter  = request.GET.get('user_id', '').strip()
    date_from       = request.GET.get('date_from', '').strip()
    date_to         = request.GET.get('date_to', '').strip()
    order_param     = request.GET.get('asc_desc', '')  # 'asc' or 'desc' from the filter

    # So the template can keep them:
    context['current_invoice_number'] = invoice_number
    context['current_owner'] = owner_id_filter
    context['current_user']  = user_id_filter
    context['current_date_from'] = date_from
    context['current_date_to']   = date_to
    context['order'] = order_param if order_param else 'desc'  # default to 'desc'

    # If user supplied ANY advanced filters, we do DB-based filtering
    has_filter = any([invoice_number, owner_id_filter, user_id_filter, date_from, date_to, order_param])

    # ---------------------------------------------
    # 5) Combine advanced GET filters if present
    # ---------------------------------------------
    qs = base_qs  # We'll apply extra filters to this
    if has_filter:
        # (A) Invoice Number
        if invoice_number:
            try:
                inv_num = int(invoice_number)
                qs = qs.filter(invoice_number=inv_num)
            except ValueError:
                qs = qs.none()  # If user typed non-integer => no results

        # (B) Owner from the Filter Form
        if owner_id_filter and owner_id_filter != 'x':
            try:
                filter_owner_id = int(owner_id_filter)
                qs = qs.filter(owner__id=filter_owner_id)
            except ValueError:
                qs = qs.none()

        # (C) Written By User
        if user_id_filter and user_id_filter != 'x':
            try:
                user_id_int = int(user_id_filter)
                qs = qs.filter(user__id=user_id_int)
            except ValueError:
                qs = qs.none()

        # (D) Date range
        # We assume "today_date" is a DateField
        if date_from and date_to:
            qs = qs.filter(today_date__range=[date_from, date_to])
        elif date_from:
            qs = qs.filter(today_date__gte=date_from)
        elif date_to:
            qs = qs.filter(today_date__lte=date_to)

        # (E) Sort by date if 'asc_desc' was provided
        if order_param == 'asc':
            qs = qs.order_by('today_date')
        else:
            qs = qs.order_by('-today_date')

    else:
        # ---------------------------------------------
        # 6) No GET filters => use your OLD POST-based sorting
        # ---------------------------------------------
        if request.method == "POST" and request.POST.get('asc_desc') in ("0", "1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                qs = qs.order_by('today_date')  # ascending
            else:
                qs = qs.order_by('-today_date') # descending
        else:
            # No sort => just keep the base queryset in default order
            # If you want default descending, do .order_by('-today_date')
            qs = qs.order_by('-today_date')

    # ---------------------------------------------
    # 7) Paginate the final qs
    # ---------------------------------------------
    paginator = Paginator(qs, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    context['page_obj'] = page_obj

    # ---------------------------------------------
    # 8) Provide owners & users for the filter form
    # ---------------------------------------------
    # If user is restricted (v/w), only show allowed owners
    if obj.type_of_user in ('v','w'):
        allowed_owners = obj.invoice_owner_allowed.all()
        all_owners = invoice_owner.objects.all()
        context['owners'] = [o for o in all_owners if o in allowed_owners]
    else:
        context['owners'] = invoice_owner.objects.all()

    context['users'] = User.objects.all()

    # Render
    return render(request, template, context)


@login_required(login_url='/')
def maintenance_invoices(request,id):
    if id:
        template = "apartments_maintenance_invoice.html"
        context = {}
        obj = get_user_profile(request.user)

        context['type_of_user'] = obj.type_of_user == 'd'
        context['write_priv'] = obj.type_of_user == 'w'
        aobj = apartment.objects.get(pk=id)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("today_date")
            else:
                objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("-today_date")
        else:
            objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj)
        context['aobj'] = aobj
        context['date_dis'] = aobj.dob.strftime("%Y-%m-%d")
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect('/home')

@login_required(login_url='/')
def maintenance_invoice_form(request,id):
    if id:
        template = "new_maintenance_invoice_form.html"
        context = {'id':id}
        objs = reversed(list(maintenance_invoice.objects.filter(is_deleted=False,apartment=apartment.objects.get(pk=id)).order_by("-today_date")))
        temp = []
        temp1 = []
        tempcount = 0
        for i in objs:
            temp.append(i.amount)
            temp1.append("On : {}".format(i.today_date))
            tempcount += 1
            if tempcount == 3:
                break
        context['prev_trans'] = zip(temp1,temp)
        if request.method == "POST":
            if request.POST['amount']:
                obj = maintenance_invoice()
                obj.user = request.user
                obj.apartment = apartment.objects.get(pk=id)
                obj.owner = apartment.objects.get(pk=id).building.owner
                obj.amount = request.POST['amount']
                if request.POST['note']:
                    obj.note = request.POST['note']
                temp_len_inv = len(maintenance_invoice.objects.filter(is_deleted=False,owner=obj.owner))
                if temp_len_inv > 0:
                    obj.invoice_number = maintenance_invoice.objects.filter(is_deleted=False,owner=obj.owner).order_by("-invoice_number")[0].invoice_number + 1
                obj.save()
                return redirect("/maintenance-invoices/{}".format(id))
            else:
                context['message'] = "Entered data is not valid."
        return render(request,template,context)
    else:
        return redirect("/home")

@login_required(login_url='/')
def delete_maintenance_invoice(request,id):
    obj = maintenance_invoice.objects.get(pk=id)
    try:
        tempId = obj.apartment.id
        obj.is_deleted = True
        obj.save()
        return redirect("/maintenance-invoices/{}".format(tempId))
    except:
        obj.is_deleted = True
        obj.save()
        return redirect('/home')

@login_required(login_url='/')
def actual_delete_maintenance_invoice(request,id):
    obj = maintenance_invoice.objects.get(pk=id)
    try:
        tempId = obj.apartment.id
        obj.delete()
        return redirect("/deleted-maintenance-invoices")
    except:
        obj.delete()
        return redirect('/home')

@login_required(login_url='/')
def owner_maintenance_invoices(request,id):
    if id:
        template = "owners_maintenance_invoice.html"
        context = {}
        obj = get_user_profile(request.user)

        context['type_of_user'] = obj.type_of_user == 'd'
        if id != "x":
            context['sel_owner'] = int(id)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                if id != "x":
                    objs = maintenance_invoice.objects.filter(is_deleted=False,owner__id=int(id)).order_by("today_date")
                else:
                    if obj.type_of_user in ('v','w'):
                        temp2 = maintenance_invoice.objects.filter(is_deleted=False).order_by("today_date")
                        objs = []
                        for i in temp2:
                            if i.owner in obj.invoice_owner_allowed.all():
                                objs.append(i)
                    else:
                        objs = maintenance_invoice.objects.filter(is_deleted=False).order_by("today_date")
            else:
                if id != "x":
                    objs = maintenance_invoice.objects.filter(is_deleted=False,owner__id=int(id)).order_by("-today_date")
                else:
                    if obj.type_of_user in ('v','w'):
                        temp2 = maintenance_invoice.objects.filter(is_deleted=False).order_by("-today_date")
                        objs = []
                        for i in temp2:
                            if i.owner in obj.invoice_owner_allowed.all():
                                objs.append(i)
                    else:
                        objs = maintenance_invoice.objects.filter(is_deleted=False).order_by("-today_date")
        else:
            if id != "x":
                objs = maintenance_invoice.objects.filter(is_deleted=False,owner__id=int(id))
            else:
                if obj.type_of_user in ('v','w'):
                    temp2 = maintenance_invoice.objects.filter(is_deleted=False)
                    objs = []
                    for i in temp2:
                        if i.owner in obj.invoice_owner_allowed.all():
                            objs.append(i)
                else:
                    objs = maintenance_invoice.objects.filter(is_deleted=False)

        context['objs'] = objs
        if obj.type_of_user in ('v','w'):
            temp2 = maintenance_invoice.objects.filter(is_deleted=False)
            temp2_list = []
            for i in temp2:
                if i in obj.invoice_owner_allowed.all():
                    temp2_list.append(i)
            context['owners'] = temp2_list
        else:
            context['owners'] = maintenance_invoice.objects.filter(is_deleted=False)
        return render(request,template,context)
    else:
        return redirect('/home')
def owner_report(request,id):
    date_col_rows = [{'row_start':4,'row_end':29,'col_start':10,'col_end':16},{'row_start':4,'row_end':29,'col_start':2,'col_end':8}]
    date_col_rows_inv = [{'row_start':4,'row_end':29,'col_start':10,'col_end':14},{'row_start':4,'row_end':29,'col_start':2,'col_end':6}]
    date_col_rows_main = [{'row_start':4,'row_end':29,'col_start':15,'col_end':16},{'row_start':4,'row_end':29,'col_start':7,'col_end':8}]
    row_offset = 31
    workbook = openpyxl.load_workbook('Monthly_reportNew.xlsx')
    worksheet = workbook.get_sheet_by_name(workbook.get_sheet_names()[0])

    from_date = request.POST['from-date'].split("-")
    from_date = datetime.date(year=int(from_date[0]),month=int(from_date[1]),day=int(from_date[2]))
    to_date = request.POST['to-date'].split("-")
    to_date = datetime.date(year=int(to_date[0]),month=int(to_date[1]),day=int(to_date[2]))

    owner_obj = invoice_owner.objects.get(pk=id)
    invoice_objs = invoice.objects.filter(is_deleted=False,owner=owner_obj,today_date__gte=from_date,today_date__lte=to_date).order_by("today_date")
    maintenance_objs = maintenance_invoice.objects.filter(is_deleted=False,owner=owner_obj,today_date__gte=from_date,today_date__lte=to_date).order_by("today_date")

    offset = 0
    for i in range(1,32):
        limits = date_col_rows[i%2]
        for r in range(limits['row_start']+offset,limits['row_end']+1+offset):
            for c in range(limits['col_start'],limits['col_end']+1):
                cell = worksheet.cell(row=r,column=c)
                cell.value = ""

        inv_data = []
        for j in invoice_objs:
            if j.today_date.day == i:
                inv_data.append(j)

        limits = date_col_rows_inv[i%2]
        for j,r in zip(inv_data,range(limits['row_start']+offset,limits['row_end']+1+offset)):
            cell = worksheet.cell(row=r,column=limits['col_start'])
            try:
                cell.value = j.apartment.aprt_number
            except:
                cell.value = " "
            cell = worksheet.cell(row=r,column=limits['col_start']+1)
            try:
                cell.value = j.apartment.building.name
            except:
                cell.value = " "
            if j.payment_method == 'Cash':
                cell = worksheet.cell(row=r,column=limits['col_start']+2)
                cell.value = j.amount
            else:
                cell = worksheet.cell(row=r,column=limits['col_start']+3)
                cell.value = j.amount
            cell = worksheet.cell(row=r,column=limits['col_start']+4)
            cell.value = j.invoice_number

        main_data = []
        for j in maintenance_objs:
            if j.today_date.day == i:
                main_data.append(j)

        limits = date_col_rows_main[i%2]
        for j,r in zip(main_data,range(limits['row_start']+offset,limits['row_end']+1+offset)):
            cell = worksheet.cell(row=r,column=limits['col_start'])
            cell.value = j.amount
            cell = worksheet.cell(row=r,column=limits['col_start']+1)
            cell.value = j.note


        if (i % 2 == 0):
            offset += row_offset

    workbook.save('Monthly_reportNew.xlsx')

    return FileResponse(open("Monthly_reportNew.xlsx","rb"), as_attachment=True)

@login_required(login_url='/')
def check_download_allowed(request,id):
    user_perms = user_profile.objects.filter(user=request.user)
    if (len(user_perms) > 0):
        owner_obj = invoice_owner.objects.get(pk=id)
        obj = user_perms[0]
        if obj.download_report_allowed:
            if owner_obj in obj.invoice_owner_allowed.all():
                return JsonResponse({'check':True})
            else:
                return JsonResponse({'check':False})
        else:
            return JsonResponse({'check':False})
    else:
        return JsonResponse({'check':True})

@login_required(login_url='/')
def check_delete_allowed(request,id,type_of):
    user_perms = user_profile.objects.filter(user=request.user)
    if (len(user_perms) > 0):
        if type_of == "m":
            obj = maintenance_invoice.objects.get(pk=id)
        else:
            obj = invoice.objects.get(pk=id)
        owner_obj = obj.owner
        obj = user_perms[0]
        if obj.delete_invoice_allowed:
            if owner_obj in obj.invoice_owner_allowed.all():
                return JsonResponse({'check':True})
            else:
                return JsonResponse({'check':False})
        else:
            return JsonResponse({'check':False})
    else:
        return JsonResponse({'check':True})

@login_required(login_url='/')
def receive_invoice(request,id):
    obj = invoice.objects.get(pk=id)
    obj.received_by = request.user.username
    obj.save()
    return redirect('/invoices/{}'.format(obj.apartment.id))

@login_required(login_url='/')
def deleted_invoices(request):
    template = "deleted_invoices.html"
    context = {}

    if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
        context["order"] = request.POST['asc_desc']
        if request.POST["asc_desc"] == "0":
            objs = invoice.objects.filter(is_deleted=True).order_by("today_date")
        else:
            objs = invoice.objects.filter(is_deleted=True).order_by("-today_date")
    else:
        objs = invoice.objects.filter(is_deleted=True)

    context['objs'] = objs
    return render(request,template,context)

@login_required(login_url='/')
def deleted_maintenance_invoices(request):
    template = "deleted_maintenance_invoices.html"
    context = {}
    if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
        context["order"] = request.POST['asc_desc']
        if request.POST["asc_desc"] == "0":
            objs = maintenance_invoice.objects.filter(is_deleted=True).order_by("today_date")
        else:
            objs = maintenance_invoice.objects.filter(is_deleted=True).order_by("-today_date")
    else:
        objs = maintenance_invoice.objects.filter(is_deleted=True)

    context['objs'] = objs
    return render(request,template,context)

@login_required(login_url='/')
def new_tenant_form(request,id,sel):
    if id:
        aobj = apartment.objects.get(pk=id)
        return redirect("/new-apartment-form/{}/{}".format(aobj.building.id,aobj.id))
    else:
        return redirect("/home")

@login_required(login_url='/')
def previous_tenants(request,id,sel):
    if id:
        template = "tenants.html"
        context = {"sel":sel}
        aobj = apartment.objects.get(pk=id)
        context['aobj'] = aobj
        try:
            context['objs'] = aobj.aprt_link.apartments.all()
        except:
            context['objs'] = []
        return render(request,template,context)
    else:
        return redirect("/home")

@login_required(login_url='/')
def tenant_invoices(request,aid,id):
    if id:
        return redirect("/invoices/{}".format(id))
        template = "tenant_invoice.html"
        context = {}
        aobj = apartment.objects.get(pk=aid)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("today_date")
            else:
                objs = invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("-today_date")
        else:
            objs = invoice.objects.filter(is_deleted=False,apartment=aobj)
        context['aobj'] = aobj
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect('/home')

@login_required(login_url='/')
def tenant_maintenance_invoices(request,aid,id):
    if id:
        return redirect("/maintenance-invoices/{}".format(id))
        template = "tenant_maintenance_invoice.html"
        context = {}
        aobj = apartment.objects.get(pk=aid)
        if request.method == "POST" and request.POST['asc_desc'] in ("0","1"):
            context["order"] = request.POST['asc_desc']
            if request.POST["asc_desc"] == "0":
                objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("today_date")
            else:
                objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj).order_by("-today_date")
        else:
            objs = maintenance_invoice.objects.filter(is_deleted=False,apartment=aobj)
        context['aobj'] = aobj
        context['objs'] = objs
        return render(request,template,context)
    else:
        return redirect('/home')

def move_up_apartment(request,bid,aid):
    aobjs = apartment.objects.filter(building__id=bid,new_tenant_added=False).order_by("display_order")
    if aobjs[0].id != aid:
        temp_order = 0
        prev_obj = aobjs[0]
        for i in aobjs:
            if i.id == aid:
                temp_order = i.display_order
                i.display_order = prev_obj.display_order
                i.save()
                prev_obj.display_order = temp_order
                prev_obj.save()
                break
            prev_obj = i
    return redirect("/apartments/{}".format(bid))

def move_down_apartment(request,bid,aid):
    aobjs = apartment.objects.filter(building__id=bid,new_tenant_added=False).order_by("-display_order")
    if aobjs[0].id != aid:
        temp_order = 0
        prev_obj = aobjs[0]
        for i in aobjs:
            if i.id == aid:
                temp_order = i.display_order
                i.display_order = prev_obj.display_order
                i.save()
                prev_obj.display_order = temp_order
                prev_obj.save()
                break
            prev_obj = i
    return redirect("/apartments/{}".format(bid))

#def search_apartment_by_phone(request):
 #   template = "apartments_by_phone_number.html"
  #  context = {}

   # context['objs'] =  apartment.objects.filter(temp_del=False,new_tenant_added=False,phone_nmber=request.POST['pnum'])

    #return render(request,template,context)

#def search_apartment_by_contract(request):
 #   template = "apartments_by_contract_number.html"
  #  context = {}

   # context['objs'] =  apartment.objects.filter(temp_del=False,new_tenant_added=False,contract_number=request.POST['cnum'])

    #return render(request,template,context)

def search_apartment_by_phone(request):
    template = "apartments_by_phone_number.html"
    context = {}

    # Use `searchInput` instead of `pnum`
    search_input = request.POST.get('searchInput')
    context['objs'] = apartment.objects.filter(temp_del=False, new_tenant_added=False, phone_nmber=search_input)

    return render(request, template, context)

def search_apartment_by_contract(request):
    template = "apartments_by_contract_number.html"
    context = {}

    # Use `searchInput` instead of `cnum`
    search_input = request.POST.get('searchInput')
    context['objs'] = apartment.objects.filter(temp_del=False, new_tenant_added=False, contract_number=search_input)

    return render(request, template, context)

#from datetime import datetime, timedelta  # Import correctly

def get_to_date_for_invoice(request, id, amt):
    retval = {"check": "0", "from_date": "", "to_date": ""}

    try:
        apt = apartment.objects.get(pk=id)
    except apartment.DoesNotExist:
        logger.error(f"Apartment with id {id} does not exist.")
        return JsonResponse(retval)

    try:
        invoices = invoice.objects.filter(is_deleted=False, apartment=apt).order_by('-id')
        for inv in invoices:
            if inv.to_date is None:
                continue

            temp_from_date = inv.to_date.strftime("%Y-%m-%d")
            temp_to_date = calculate_to_date(int(inv.apartment.annual_rent), int(amt), temp_from_date)

            retval['check'] = "1"
            retval['from_date'] = temp_from_date
            retval['to_date'] = temp_to_date
            break
    except Exception as e:
        logger.error(f"Error while processing invoices for apartment {id}: {str(e)}")

    return JsonResponse(retval)

def calculate_to_date(annual_rent, amount, from_date_str):
    from_date = datetime.datetime.strptime(from_date_str, "%Y-%m-%d")
    from_date += datetime.timedelta(days=1)

    if annual_rent == 0:
        return from_date.strftime("%Y-%m-%d")

    daily_rent = annual_rent / 365
    days_covered = amount / daily_rent
    to_date = from_date + datetime.timedelta(days=int(days_covered))

    return to_date.strftime("%Y-%m-%d")




@login_required(login_url='/')
def requests_view(request):
    return render(request, 'requests.html')














